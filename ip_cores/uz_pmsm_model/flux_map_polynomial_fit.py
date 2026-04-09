from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


NUM_GRID_POINTS = 20
MILLI_TO_BASE = 1e-3


@dataclass(frozen=True)
class FluxMap:
    id_grid: np.ndarray
    iq_grid: np.ndarray
    psi_d: np.ndarray
    psi_q: np.ndarray


def _find_row(ws, label: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    raise ValueError(f"Could not find row label {label!r} in worksheet {ws.title!r}")


def _read_numeric_block(ws, start_row: int, rows: int = NUM_GRID_POINTS, cols: int = NUM_GRID_POINTS) -> np.ndarray:
    values = []
    for row in range(start_row, start_row + rows):
        row_values = []
        for column in range(1, cols + 1):
            value = ws.cell(row=row, column=column).value
            if value is None:
                raise ValueError(f"Encountered empty cell at row {row}, column {column}")
            row_values.append(float(value))
        values.append(row_values)
    return np.asarray(values, dtype=float)


def load_flux_map(workbook_path: Path) -> FluxMap:
    worksheet = load_workbook(workbook_path, data_only=True).active

    d_current_row = _find_row(worksheet, "d-Current in A (x-Axis)")
    q_current_row = _find_row(worksheet, "q-Current in A (y-Axis)")
    d_flux_row = _find_row(worksheet, "d-Flux Values in mVs")
    q_flux_row = _find_row(worksheet, "q-Flux Values in mVs")

    id_grid_d = _read_numeric_block(worksheet, d_current_row + 1)
    iq_grid_d = _read_numeric_block(worksheet, q_current_row + 1)
    psi_d = _read_numeric_block(worksheet, d_flux_row + 1) * MILLI_TO_BASE

    q_axis_anchor = _find_row(worksheet, "FluxMap in q-Axis")
    id_grid_q = _read_numeric_block(worksheet, q_axis_anchor + 2)
    iq_grid_q = _read_numeric_block(worksheet, q_axis_anchor + 23)
    psi_q = _read_numeric_block(worksheet, q_flux_row + 1) * MILLI_TO_BASE

    if not np.allclose(id_grid_d, id_grid_q):
        raise ValueError("The d-axis and q-axis d-current grids do not match")
    if not np.allclose(iq_grid_d, iq_grid_q):
        raise ValueError("The d-axis and q-axis q-current grids do not match")

    return FluxMap(id_grid=id_grid_d, iq_grid=iq_grid_d, psi_d=psi_d, psi_q=psi_q)


def build_design_matrix(id_grid: np.ndarray, iq_grid: np.ndarray) -> np.ndarray:
    id_values = id_grid.ravel()
    iq_values = iq_grid.ravel()
    return np.column_stack(
        [
            id_values,
            iq_values,
            id_values * id_values,
            iq_values * iq_values,
            id_values * iq_values,
            np.ones_like(id_values),
        ]
    )


def fit_quadratic_surface(id_grid: np.ndarray, iq_grid: np.ndarray, psi: np.ndarray) -> np.ndarray:
    design_matrix = build_design_matrix(id_grid, iq_grid)
    coefficients, _, _, _ = np.linalg.lstsq(design_matrix, psi.ravel(), rcond=None)
    return coefficients


def evaluate_quadratic_surface(id_grid: np.ndarray, iq_grid: np.ndarray, coefficients: np.ndarray) -> np.ndarray:
    a_coeff, b_coeff, c_coeff, d_coeff, e_coeff, g_coeff = coefficients
    return (
        a_coeff * id_grid
        + b_coeff * iq_grid
        + c_coeff * id_grid * id_grid
        + d_coeff * iq_grid * iq_grid
        + e_coeff * id_grid * iq_grid
        + g_coeff
    )


def summarize_fit(label: str, coefficients: np.ndarray, measured: np.ndarray, fitted: np.ndarray) -> dict[str, float]:
    rmse = float(np.sqrt(np.mean((measured - fitted) ** 2)))
    max_abs_error = float(np.max(np.abs(measured - fitted)))

    if label == "psi_d":
        names = ["a_d", "b_d", "c_d", "d_d", "e_d", "g_d"]
    else:
        names = ["a_q", "b_q", "c_q", "d_q", "e_q", "g_q"]

    result = {name: float(value) for name, value in zip(names, coefficients, strict=True)}
    result["rmse"] = rmse
    result["max_abs_error"] = max_abs_error
    return result


def plot_flux_maps(flux_map: FluxMap, psi_d_fit: np.ndarray, psi_q_fit: np.ndarray, output_path: Path | None) -> None:
    figure = plt.figure(figsize=(18, 10))

    axes = [
        figure.add_subplot(2, 3, 1, projection="3d"),
        figure.add_subplot(2, 3, 2, projection="3d"),
        figure.add_subplot(2, 3, 3),
        figure.add_subplot(2, 3, 4, projection="3d"),
        figure.add_subplot(2, 3, 5, projection="3d"),
        figure.add_subplot(2, 3, 6),
    ]

    axes[0].plot_surface(flux_map.id_grid, flux_map.iq_grid, flux_map.psi_d, cmap="viridis")
    axes[0].set_title("Measured psi_d")
    axes[1].plot_surface(flux_map.id_grid, flux_map.iq_grid, psi_d_fit, cmap="plasma")
    axes[1].set_title("Quadratic fit psi_d")

    psi_d_error = np.abs(flux_map.psi_d - psi_d_fit)
    psi_d_error_plot = axes[2].contourf(flux_map.id_grid, flux_map.iq_grid, psi_d_error, levels=20, cmap="magma")
    axes[2].set_title("Absolute error psi_d")
    figure.colorbar(psi_d_error_plot, ax=axes[2], shrink=0.85)

    axes[3].plot_surface(flux_map.id_grid, flux_map.iq_grid, flux_map.psi_q, cmap="viridis")
    axes[3].set_title("Measured psi_q")
    axes[4].plot_surface(flux_map.id_grid, flux_map.iq_grid, psi_q_fit, cmap="plasma")
    axes[4].set_title("Quadratic fit psi_q")

    psi_q_error = np.abs(flux_map.psi_q - psi_q_fit)
    psi_q_error_plot = axes[5].contourf(flux_map.id_grid, flux_map.iq_grid, psi_q_error, levels=20, cmap="magma")
    axes[5].set_title("Absolute error psi_q")
    figure.colorbar(psi_q_error_plot, ax=axes[5], shrink=0.85)

    for axis in axes[:2] + axes[3:5]:
        axis.set_xlabel("i_d [A]")
        axis.set_ylabel("i_q [A]")
        axis.set_zlabel("psi [Vs]")

    for axis in [axes[2], axes[5]]:
        axis.set_xlabel("i_d [A]")
        axis.set_ylabel("i_q [A]")

    figure.suptitle("Flux-map polynomial approximation")
    figure.tight_layout()

    overlay_figure = plt.figure(figsize=(14, 6))
    overlay_axes = [
        overlay_figure.add_subplot(1, 2, 1, projection="3d"),
        overlay_figure.add_subplot(1, 2, 2, projection="3d"),
    ]

    overlay_axes[0].plot_surface(
        flux_map.id_grid,
        flux_map.iq_grid,
        flux_map.psi_d,
        cmap="viridis",
        alpha=0.75,
        edgecolor="none",
    )
    overlay_axes[0].plot_surface(
        flux_map.id_grid,
        flux_map.iq_grid,
        psi_d_fit,
        cmap="plasma",
        alpha=0.45,
        edgecolor="none",
    )
    overlay_axes[0].set_title("Measured and fitted psi_d")

    overlay_axes[1].plot_surface(
        flux_map.id_grid,
        flux_map.iq_grid,
        flux_map.psi_q,
        cmap="viridis",
        alpha=0.75,
        edgecolor="none",
    )
    overlay_axes[1].plot_surface(
        flux_map.id_grid,
        flux_map.iq_grid,
        psi_q_fit,
        cmap="plasma",
        alpha=0.45,
        edgecolor="none",
    )
    overlay_axes[1].set_title("Measured and fitted psi_q")

    for axis in overlay_axes:
        axis.set_xlabel("i_d [A]")
        axis.set_ylabel("i_q [A]")
        axis.set_zlabel("psi [Vs]")

    overlay_figure.suptitle("Measured versus fitted flux surfaces")
    overlay_figure.tight_layout()

    if output_path is not None:
        figure.savefig(output_path, dpi=200, bbox_inches="tight")
        overlay_path = output_path.with_name(f"{output_path.stem}_overlay{output_path.suffix}")
        overlay_figure.savefig(overlay_path, dpi=200, bbox_inches="tight")
    else:
        plt.show()


def parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    default_workbook = script_dir / "FluxMapData_Prototyp_1000rpm.xlsx"
    default_plot = script_dir / "flux_map_polynomial_fit.png"
    default_coefficients = script_dir / "flux_map_polynomial_coefficients.json"

    parser = argparse.ArgumentParser(
        description="Read a PMSM flux-map workbook, fit quadratic surfaces for psi_d and psi_q, and plot the result."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        default=default_workbook,
        help="Path to the flux-map workbook (.xlsx)",
    )
    parser.add_argument(
        "--save-plot",
        action="store_true",
        help="Save the plot to a PNG instead of opening an interactive window.",
    )
    parser.add_argument(
        "--plot-path",
        type=Path,
        default=default_plot,
        help="Target path for the saved PNG when --save-plot is used.",
    )
    parser.add_argument(
        "--coefficients-path",
        type=Path,
        default=default_coefficients,
        help="Target path for the fitted coefficients as JSON.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    flux_map = load_flux_map(args.workbook)

    psi_d_coefficients = fit_quadratic_surface(flux_map.id_grid, flux_map.iq_grid, flux_map.psi_d)
    psi_q_coefficients = fit_quadratic_surface(flux_map.id_grid, flux_map.iq_grid, flux_map.psi_q)

    psi_d_fit = evaluate_quadratic_surface(flux_map.id_grid, flux_map.iq_grid, psi_d_coefficients)
    psi_q_fit = evaluate_quadratic_surface(flux_map.id_grid, flux_map.iq_grid, psi_q_coefficients)

    coefficient_summary = {
        "psi_d": summarize_fit("psi_d", psi_d_coefficients, flux_map.psi_d, psi_d_fit),
        "psi_q": summarize_fit("psi_q", psi_q_coefficients, flux_map.psi_q, psi_q_fit),
    }

    print(json.dumps(coefficient_summary, indent=2, sort_keys=False))
    args.coefficients_path.write_text(json.dumps(coefficient_summary, indent=2), encoding="utf-8")

    plot_path = args.plot_path if args.save_plot else None
    plot_flux_maps(flux_map, psi_d_fit, psi_q_fit, plot_path)


if __name__ == "__main__":
    main()